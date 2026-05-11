"""Descarga el catalogo SAT c_ClaveProdServ y genera sat_catalog.json.

USO (corre una sola vez):

  cd backend
  python scripts/build_sat_catalog.py

  Esto:
  1. Descarga catCFDI_V_4_23032023.xls del SAT (~10MB)
  2. Extrae la hoja c_ClaveProdServ (~58k entradas)
  3. Guarda app/data/sat_catalog.json (~5MB)
  4. Commit el JSON al repo para que Render lo tenga al desplegar

Si la URL del SAT cambia, edita SAT_URL abajo o pasa la ruta local:
  python scripts/build_sat_catalog.py --local archivo.xls
"""
import argparse
import json
import sys
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).parent.parent
OUTPUT = ROOT / "app" / "data" / "sat_catalog.json"
LOCAL_DEFAULT = ROOT / "scripts" / "sat_input.xlsx"
LOCAL_XLS = ROOT / "scripts" / "sat_input.xls"

# URL oficial del SAT con todos los catalogos CFDI 4.0 (sheet c_ClaveProdServ)
SAT_URL = "http://omawww.sat.gob.mx/tramitesyservicios/Paginas/documentos/catCFDI_V_4_23032023.xls"


def descargar():
    try:
        import httpx
    except ImportError:
        print("Necesitas httpx. Instalalo: pip install httpx")
        sys.exit(1)
    print(f"Descargando catalogo SAT...")
    print(f"  URL: {SAT_URL}")
    try:
        r = httpx.get(SAT_URL, timeout=120, follow_redirects=True)
        r.raise_for_status()
        print(f"  Descargado: {len(r.content) / 1024:.0f} KB")
        return r.content
    except Exception as e:
        print(f"ERROR al descargar: {e}")
        print(f"Alternativa: descargalo manualmente y pasalo con --local")
        sys.exit(1)


def procesar(xls_bytes: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Necesitas openpyxl. Instalalo: pip install openpyxl")
        sys.exit(1)

    # openpyxl no lee .xls (formato viejo BIFF), solo .xlsx
    # Si bytes empiezan con D0CF (signature de XLS viejo), intentamos xlrd
    if xls_bytes[:2] == b"\xd0\xcf":
        print("Archivo es .xls (formato antiguo). Intentando con xlrd...")
        try:
            import xlrd
        except ImportError:
            print("ERROR: necesitas xlrd para archivos .xls antiguos")
            print("Instala con: pip install xlrd==1.2.0")
            sys.exit(1)
        wb = xlrd.open_workbook(file_contents=xls_bytes)
        target = None
        for name in wb.sheet_names():
            if "ProdServ" in name or "claveprodserv" in name.lower():
                target = name
                break
        if not target:
            print(f"Hojas disponibles: {wb.sheet_names()}")
            sys.exit(1)
        ws = wb.sheet_by_name(target)
        rows = ((ws.cell_value(r, c) for c in range(ws.ncols)) for r in range(ws.nrows))
        rows_list = [list(r) for r in rows]
    else:
        wb = load_workbook(BytesIO(xls_bytes), read_only=True, data_only=True)
        target = None
        for name in wb.sheetnames:
            if "ProdServ" in name or "claveprodserv" in name.lower():
                target = name
                break
        if not target:
            print(f"Hojas disponibles: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[target]
        rows_list = [list(r) for r in ws.iter_rows(values_only=True)]

    print(f"Procesando hoja '{target}' con {len(rows_list)} filas...")

    # Debug: muestra primeras 3 filas para verificar formato
    print("Muestra primeras 3 filas:")
    for i, r in enumerate(rows_list[:3]):
        print(f"  Fila {i}: {r[:4] if r else '(vacia)'}")

    def normalizar_clave(val):
        """Convierte cualquier valor a clave de 8 digitos o devuelve None."""
        if val is None:
            return None
        # xlrd lee numericos como float
        if isinstance(val, float):
            val = int(val)
        if isinstance(val, int):
            val = str(val)
        else:
            val = str(val).strip()
        # Solo digitos
        if not val.isdigit():
            return None
        # Padding con ceros si es 7 (algunos catalogos pierden el cero inicial)
        if len(val) == 7:
            val = "0" + val
        return val if len(val) == 8 else None

    catalog = []
    for row in rows_list:
        if not row:
            continue
        clave = normalizar_clave(row[0])
        if not clave:
            continue
        # Descripcion suele estar en columna 1 (B) o 2 (C)
        desc = ""
        for i in range(1, min(5, len(row))):
            val = row[i]
            if val is None:
                continue
            s = str(val).strip()
            if len(s) > 3 and not s.replace("-", "").replace("/", "").isdigit():
                desc = s
                break
        if desc:
            catalog.append({"clave": clave, "descripcion": desc})

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(catalog, ensure_ascii=False, separators=(",", ":")))
    print(f"OK: {len(catalog)} claves SAT guardadas")
    print(f"  Archivo: {OUTPUT}")
    print(f"  Tamano:  {OUTPUT.stat().st_size / 1024:.0f} KB")
    print()
    print("Siguiente paso:")
    print("  cd ..")
    print("  git add backend/app/data/sat_catalog.json")
    print("  git commit -m 'data: catalogo SAT c_ClaveProdServ completo'")
    print("  git push")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--local", help="Ruta a archivo .xls/.xlsx local en lugar de descargar del SAT")
    args = ap.parse_args()

    if args.local:
        path = Path(args.local)
        if not path.exists():
            print(f"ERROR: archivo no existe: {path}")
            sys.exit(1)
        xls_bytes = path.read_bytes()
    elif LOCAL_DEFAULT.exists():
        print(f"Usando archivo local: {LOCAL_DEFAULT}")
        xls_bytes = LOCAL_DEFAULT.read_bytes()
    elif LOCAL_XLS.exists():
        print(f"Usando archivo local: {LOCAL_XLS}")
        xls_bytes = LOCAL_XLS.read_bytes()
    else:
        xls_bytes = descargar()

    procesar(xls_bytes)


if __name__ == "__main__":
    main()
