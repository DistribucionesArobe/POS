"""Parseo de cotizaciones desde XLSX o imagen/PDF para pre-llenar la Caja.

Flujo:
1. Detecta tipo de archivo por extension/mime
2. XLSX  -> openpyxl
   Imagen -> Claude Vision
   PDF   -> primera pagina como imagen, Claude Vision
3. Hace fuzzy match de cada linea contra el catalogo de VarianteProducto
4. Devuelve lista de lineas con match_variante_id (o None si no encuentra)
"""
from __future__ import annotations
from difflib import SequenceMatcher
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Producto, VarianteProducto


def parsear_xlsx(file_bytes: bytes) -> list[dict]:
    """Lee columnas Descripcion / Unidad / Cantidad / Precio / Monto de una cotizacion."""
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Detecta fila de headers (suele estar entre las primeras 5)
    headers = None
    header_row_idx = 1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        vals = [str(c or "").strip().lower() for c in row]
        if any("descrip" in v for v in vals) and any("cantid" in v or "cant" == v for v in vals):
            headers = vals
            header_row_idx = i
            break

    if not headers:
        raise ValueError("No encontre los headers Descripcion/Cantidad en las primeras 10 filas")

    def find_col(*kws):
        for kw in kws:
            for i, h in enumerate(headers):
                if kw in h:
                    return i
        return None

    idx_desc = find_col("descrip")
    idx_unidad = find_col("unidad", "u. de m", "u/m")
    idx_cant = find_col("cantid", "cant")
    idx_precio = find_col("precio")
    idx_monto = find_col("monto", "importe", "total")

    if idx_desc is None or idx_cant is None:
        raise ValueError("Faltan columnas obligatorias: Descripcion y Cantidad")

    lineas = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(c is None for c in row):
            continue
        desc = row[idx_desc] if idx_desc < len(row) else None
        if not desc or not str(desc).strip():
            continue
        desc_str = str(desc).strip()
        # Skipear filas de subtotal/iva/total
        if desc_str.lower() in ("subtotal", "iva", "total", "total de material"):
            continue
        try:
            cantidad = float(row[idx_cant] or 0) if idx_cant < len(row) else 0
        except (TypeError, ValueError):
            cantidad = 0
        if cantidad <= 0:
            continue
        try:
            precio = float(row[idx_precio] or 0) if idx_precio is not None and idx_precio < len(row) else 0
        except (TypeError, ValueError):
            precio = 0
        try:
            monto = float(row[idx_monto] or 0) if idx_monto is not None and idx_monto < len(row) else cantidad * precio
        except (TypeError, ValueError):
            monto = cantidad * precio
        unidad = ""
        if idx_unidad is not None and idx_unidad < len(row) and row[idx_unidad]:
            unidad = str(row[idx_unidad]).strip()
        lineas.append({
            "descripcion": desc_str,
            "unidad": unidad,
            "cantidad": cantidad,
            "precio": precio,
            "monto": monto,
        })
    return lineas


def parsear_imagen(file_bytes: bytes, mime_type: str) -> list[dict]:
    """Usa Claude Vision para extraer lineas de una imagen/PDF."""
    from app.integrations.anthropic_client import ClaudeClient
    client = ClaudeClient()
    raw = client.parsear_cotizacion_imagen(file_bytes, mime_type=mime_type)
    # Normalizar
    out = []
    for item in (raw or []):
        if not item.get("descripcion"):
            continue
        try:
            cantidad = float(item.get("cantidad") or 0)
        except (TypeError, ValueError):
            cantidad = 0
        if cantidad <= 0:
            continue
        try:
            precio = float(item.get("precio") or 0)
        except (TypeError, ValueError):
            precio = 0
        try:
            monto = float(item.get("monto") or 0)
        except (TypeError, ValueError):
            monto = cantidad * precio
        out.append({
            "descripcion": str(item["descripcion"]).strip(),
            "unidad": str(item.get("unidad") or "").strip(),
            "cantidad": cantidad,
            "precio": precio,
            "monto": monto or cantidad * precio,
        })
    return out


_STOPWORDS = {"de", "del", "para", "con", "sin", "la", "el", "los", "las",
              "y", "o", "default", "pza", "pieza", "kg", "lt", "m", "cm",
              "ml", "incluye"}


def _tokens(s: str) -> set[str]:
    """Tokeniza ignorando palabras comunes y caracteres no alfanumericos."""
    import re
    s = s.lower()
    # Reemplaza separadores comunes con espacio
    s = re.sub(r"[^a-z0-9áéíóúñ]+", " ", s)
    tokens = {t for t in s.split() if len(t) > 1 and t not in _STOPWORDS}
    return tokens


def _similitud(a: str, b: str) -> float:
    """Combina similitud de tokens (Jaccard) con similitud de caracteres."""
    # Jaccard sobre tokens distintivos
    ta = _tokens(a)
    tb = _tokens(b)
    if ta and tb:
        inter = len(ta & tb)
        union = len(ta | tb)
        jaccard = inter / union if union else 0.0
    else:
        jaccard = 0.0
    # Bonus: si los tokens distintivos coinciden 100%, sube agresivo
    if ta and ta.issubset(tb):
        jaccard = max(jaccard, 0.85)
    # Promedio con SequenceMatcher como tiebreaker
    char_sim = SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()
    return jaccard * 0.7 + char_sim * 0.3


def matchear_lineas(db: Session, empresa_id: int, lineas: list[dict],
                     umbral: float = 0.75) -> list[dict]:
    """Para cada linea busca la mejor variante en el catalogo (por nombre + presentacion).

    Umbral 0.75 = match conservador. Bajalo si quieres mas matches automaticos
    (a riesgo de falsos positivos). Subelo si quieres certeza casi total.

    Devuelve la misma lista enriquecida con:
      - match_variante_id (int|None)
      - match_score (0.0-1.0)
      - match_nombre (string|None)
      - match_sku (string|None)
      - match_precio_catalogo (float|None)
    """
    # Cargar todas las variantes de la empresa una sola vez
    variantes = (
        db.query(VarianteProducto, Producto)
        .join(Producto, Producto.id == VarianteProducto.producto_id)
        .filter(Producto.empresa_id == empresa_id)
        .filter(VarianteProducto.activo == True)  # noqa
        .all()
    )
    catalogo = [
        {
            "id": v.id, "sku": v.sku,
            "nombre": f"{p.nombre} {v.presentacion}".strip(),
            "precio": float(v.precio_publico or 0),
            "stock": float(v.stock_actual or 0),
            "tasa_iva": float(v.tasa_iva) if v.tasa_iva is not None else 0.16,
        }
        for v, p in variantes
    ]

    resultado = []
    for linea in lineas:
        descripcion = linea["descripcion"]
        mejor = None
        mejor_score = 0.0
        for c in catalogo:
            score = _similitud(descripcion, c["nombre"])
            if score > mejor_score:
                mejor_score = score
                mejor = c
        if mejor and mejor_score >= umbral:
            resultado.append({
                **linea,
                "match_variante_id": mejor["id"],
                "match_score": round(mejor_score, 2),
                "match_nombre": mejor["nombre"],
                "match_sku": mejor["sku"],
                "match_precio_catalogo": mejor["precio"],
                "match_stock": mejor["stock"],
                "match_tasa_iva": mejor["tasa_iva"],
            })
        else:
            resultado.append({
                **linea,
                "match_variante_id": None,
                "match_score": round(mejor_score, 2) if mejor else 0.0,
                "match_nombre": None,
                "match_sku": None,
                "match_precio_catalogo": None,
                "match_stock": None,
                "match_tasa_iva": None,
            })
    return resultado
