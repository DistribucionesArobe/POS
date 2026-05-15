"""Importacion de productos desde Excel (.xlsx)."""
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models import Producto, VarianteProducto
from app.services import inventario_service


COLUMNAS = [
    ("nombre", "Nombre del producto *"),
    ("categoria", "Familia / Categoria"),
    ("marca", "Marca"),
    ("sku", "SKU *"),
    ("presentacion", "Presentacion"),
    ("unidad", "Unidad (PZA, KG, M, BULTO)"),
    ("precio", "Precio publico *"),
    ("costo", "Costo"),
    ("stock_inicial", "Stock inicial"),
    ("stock_minimo", "Stock minimo"),
    ("clave_sat", "Clave SAT (8 digitos)"),
]


def generar_plantilla() -> bytes:
    """Genera plantilla XLSX vacia con headers + 1 fila ejemplo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1f2937")
    for col_idx, (key, label) in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22

    # Fila ejemplo
    ejemplo = [
        "Varilla corrugada 3/8", "Aceros", "Acerex",
        "VAR-3-8-12M", "Entera 12m", "PZA",
        180.00, 140.00, 50, 5, "30102404",
    ]
    for col_idx, val in enumerate(ejemplo, 1):
        ws.cell(row=2, column=col_idx, value=val)

    # Nota
    ws.cell(row=4, column=1, value="* obligatorios. SKU debe ser unico por empresa.").font = Font(italic=True, color="6b7280")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_productos(db: Session, file_bytes: bytes, empresa_id: int) -> dict:
    """Lee XLSX, crea productos+variantes. Devuelve resumen."""
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo: {e}")

    # Mapear columnas por header
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip().lower())

    def find_col(*names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return None

    idx_nombre = find_col("nombre")
    idx_categoria = find_col("familia", "categoria")
    idx_marca = find_col("marca")
    idx_sku = find_col("sku")
    idx_presentacion = find_col("presentacion")
    idx_unidad = find_col("unidad")
    idx_precio = find_col("precio")
    idx_costo = find_col("costo")
    idx_stock = find_col("stock inicial", "stock_inicial")
    idx_min = find_col("stock minimo", "stock_minimo")
    idx_sat = find_col("clave sat", "clave_sat")

    if idx_nombre is None or idx_sku is None or idx_precio is None:
        raise ValueError(
            "El archivo debe tener columnas: 'Nombre', 'SKU' y 'Precio' (revisa los headers en la fila 1)"
        )

    creados = 0
    actualizados = 0
    errores: list[str] = []
    # Commit por lotes para evitar OOM / timeout en uploads grandes (5k+ filas).
    # Cada 300 filas vaciamos la sesion al DB. Antes el commit unico al final con
    # 5000 productos + 5000 movimientos kardex se pasaba de RAM en Render starter.
    BATCH = 300

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            if not row or not row[idx_nombre] or not row[idx_sku]:
                continue
            nombre = str(row[idx_nombre]).strip()
            sku = str(row[idx_sku]).strip()
            precio = float(row[idx_precio] or 0)
            costo = float(row[idx_costo] or 0) if idx_costo is not None else 0
            stock_inicial = float(row[idx_stock] or 0) if idx_stock is not None else 0
            stock_min = float(row[idx_min] or 0) if idx_min is not None else 0
            categoria = str(row[idx_categoria]).strip() if idx_categoria is not None and row[idx_categoria] else None
            marca = str(row[idx_marca]).strip() if idx_marca is not None and row[idx_marca] else None
            presentacion = str(row[idx_presentacion]).strip() if idx_presentacion is not None and row[idx_presentacion] else "Default"
            unidad = str(row[idx_unidad]).strip().upper() if idx_unidad is not None and row[idx_unidad] else "PZA"
            clave_sat = str(row[idx_sat]).strip() if idx_sat is not None and row[idx_sat] else None

            # Si SKU ya existe, actualiza precio/costo en lugar de crear
            existente = db.query(VarianteProducto).filter(VarianteProducto.sku == sku).first()
            if existente:
                producto = db.get(Producto, existente.producto_id)
                if producto.empresa_id != empresa_id:
                    errores.append(f"Fila {row_idx}: SKU '{sku}' pertenece a otra empresa")
                    continue
                existente.precio_publico = precio
                if costo:
                    existente.costo_promedio = costo
                if stock_min:
                    existente.stock_minimo = stock_min
                actualizados += 1
                continue

            # Crear producto + variante
            # Si la categoria existe, reusar el producto (para no duplicar familias)
            producto = None
            if categoria:
                producto = (
                    db.query(Producto)
                    .filter(Producto.empresa_id == empresa_id)
                    .filter(Producto.nombre == nombre)
                    .filter(Producto.categoria == categoria)
                    .first()
                )
            if not producto:
                producto = Producto(
                    empresa_id=empresa_id, nombre=nombre, categoria=categoria,
                    marca=marca, clave_prod_serv_sat=clave_sat,
                )
                db.add(producto)
                db.flush()

            variante = VarianteProducto(
                producto_id=producto.id, sku=sku, presentacion=presentacion,
                unidad=unidad, precio_publico=precio, costo_promedio=costo,
                stock_minimo=stock_min,
            )
            db.add(variante)
            db.flush()

            # Stock inicial via kardex
            if stock_inicial > 0:
                inventario_service.aplicar_movimiento(
                    db, variante.id, "ENTRADA_COMPRA", stock_inicial,
                    empresa_id=empresa_id,
                    costo_unitario=costo,
                    notas="Carga inicial por importacion Excel",
                )

            creados += 1
        except Exception as e:
            errores.append(f"Fila {row_idx}: {e}")
            # Si rompe la transaccion, hay que rollback y limpiar el batch.
            try:
                db.rollback()
            except Exception:
                pass

        # Commit por lotes
        if (creados + actualizados) > 0 and (creados + actualizados) % BATCH == 0:
            try:
                db.commit()
                # Limpia identity map para liberar RAM (los objetos siguen en DB)
                db.expunge_all()
            except Exception as e:
                errores.append(f"Lote alrededor de fila {row_idx}: {e}")
                db.rollback()

    # Commit final del batch parcial
    try:
        db.commit()
    except Exception as e:
        errores.append(f"Commit final: {e}")
        db.rollback()

    return {
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores,
    }
