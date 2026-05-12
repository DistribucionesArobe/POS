"""Reportes - filtrados por empresa."""
from datetime import date, datetime, timedelta
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.db import get_db
from app.models import (
    DocumentoVenta, ConceptoVenta, CuentaPorCobrar, VarianteProducto,
    Cliente, Producto, Pago, CorteCaja, Usuario,
)
from app.models.venta import EstatusDocumento, TipoDocumento
from app.services.security import get_active_empresa_id, get_current_user

router = APIRouter()


# ---- Etiquetas de formas de pago SAT ----
FORMA_LABEL = {
    "01": "Efectivo", "02": "Cheque", "03": "Transferencia",
    "04": "Tarjeta crédito", "28": "Tarjeta débito", "99": "Por definir",
}


def _xlsx_response(wb, filename: str) -> Response:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _rango(periodo: str) -> tuple[datetime, datetime]:
    """Devuelve (inicio, fin) UTC para el periodo dado."""
    today = datetime.utcnow().date()
    if periodo == "hoy":
        inicio = datetime.combine(today, datetime.min.time())
        return inicio, inicio + timedelta(days=1)
    if periodo == "semana":
        lunes = today - timedelta(days=today.weekday())
        inicio = datetime.combine(lunes, datetime.min.time())
        return inicio, inicio + timedelta(days=7)
    if periodo == "mes":
        primero = today.replace(day=1)
        inicio = datetime.combine(primero, datetime.min.time())
        # primer dia del mes siguiente
        if primero.month == 12:
            sig = primero.replace(year=primero.year + 1, month=1)
        else:
            sig = primero.replace(month=primero.month + 1)
        return inicio, datetime.combine(sig, datetime.min.time())
    # default: hoy
    inicio = datetime.combine(today, datetime.min.time())
    return inicio, inicio + timedelta(days=1)


@router.get("/dashboard")
def dashboard(
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """KPIs + ventas por periodo + top productos + comparativa vs periodo anterior."""

    def _resumen(ini: datetime, fin: datetime) -> dict:
        q = db.query(
            func.coalesce(func.sum(DocumentoVenta.total), 0).label("total"),
            func.count().label("n"),
            func.coalesce(func.avg(DocumentoVenta.total), 0).label("avg"),
        ).filter(
            DocumentoVenta.empresa_id == empresa_id,
            DocumentoVenta.fecha >= ini,
            DocumentoVenta.fecha < fin,
            DocumentoVenta.tipo.in_([
                TipoDocumento.TICKET.value, TipoDocumento.FACTURA.value,
                TipoDocumento.REMISION.value,
            ]),
            DocumentoVenta.estatus != EstatusDocumento.CANCELADO.value,
        ).one()
        return {"total": float(q.total or 0), "n": q.n, "ticket_promedio": float(q.avg or 0)}

    periodos: dict[str, dict] = {}
    for p in ("hoy", "semana", "mes"):
        ini, fin = _rango(p)
        actual = _resumen(ini, fin)
        # Periodo anterior del mismo largo
        delta = fin - ini
        prev = _resumen(ini - delta, ini)
        if prev["total"] > 0:
            cambio_pct = (actual["total"] - prev["total"]) / prev["total"] * 100
        else:
            cambio_pct = None
        periodos[p] = {**actual, "vs_anterior": prev, "cambio_pct": cambio_pct}

    # Top productos del mes por monto
    ini_mes, fin_mes = _rango("mes")
    top = (
        db.query(
            ConceptoVenta.descripcion,
            func.sum(ConceptoVenta.cantidad).label("cant"),
            func.sum(ConceptoVenta.importe).label("monto"),
        )
        .join(DocumentoVenta, DocumentoVenta.id == ConceptoVenta.documento_id)
        .filter(
            DocumentoVenta.empresa_id == empresa_id,
            DocumentoVenta.fecha >= ini_mes,
            DocumentoVenta.fecha < fin_mes,
            DocumentoVenta.tipo.in_([
                TipoDocumento.TICKET.value, TipoDocumento.FACTURA.value,
                TipoDocumento.REMISION.value,
            ]),
            DocumentoVenta.estatus != EstatusDocumento.CANCELADO.value,
        )
        .group_by(ConceptoVenta.descripcion)
        .order_by(func.sum(ConceptoVenta.importe).desc())
        .limit(10)
        .all()
    )

    # Clientes nuevos del mes
    nuevos = db.query(Cliente).filter(
        Cliente.empresa_id == empresa_id,
        Cliente.creado_en >= ini_mes,
        Cliente.creado_en < fin_mes,
    ).count()

    # Cartera total pendiente
    cartera = (
        db.query(func.coalesce(func.sum(CuentaPorCobrar.saldo), 0))
        .join(DocumentoVenta, DocumentoVenta.id == CuentaPorCobrar.documento_id)
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(CuentaPorCobrar.pagado == False)
        .scalar()
    )

    return {
        "periodos": periodos,
        "top_productos": [
            {"descripcion": d, "cantidad": float(c or 0), "monto": float(m or 0)}
            for d, c, m in top
        ],
        "clientes_nuevos_mes": nuevos,
        "cartera_total": float(cartera or 0),
    }



@router.get("/kpis")
def kpis(
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    today = datetime.utcnow().date()
    inicio = datetime.combine(today, datetime.min.time())
    fin = inicio + timedelta(days=1)

    productos_stock = (
        db.query(VarianteProducto)
        .join(Producto)
        .filter(Producto.empresa_id == empresa_id)
        .filter(VarianteProducto.activo == True, VarianteProducto.stock_actual > 0)
        .count()
    )

    ventas_hoy = db.query(
        func.coalesce(func.sum(DocumentoVenta.total), 0)
    ).filter(
        DocumentoVenta.empresa_id == empresa_id,
        DocumentoVenta.fecha >= inicio,
        DocumentoVenta.fecha < fin,
        DocumentoVenta.estatus != EstatusDocumento.CANCELADO.value,
    ).scalar()

    docs_hoy = db.query(DocumentoVenta).filter(
        DocumentoVenta.empresa_id == empresa_id,
        DocumentoVenta.fecha >= inicio,
        DocumentoVenta.fecha < fin,
        DocumentoVenta.estatus != EstatusDocumento.CANCELADO.value,
    ).count()

    cartera = (
        db.query(func.coalesce(func.sum(CuentaPorCobrar.saldo), 0))
        .join(DocumentoVenta, DocumentoVenta.id == CuentaPorCobrar.documento_id)
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(CuentaPorCobrar.pagado == False)
        .scalar()
    )

    docs_pendientes = (
        db.query(CuentaPorCobrar)
        .join(DocumentoVenta, DocumentoVenta.id == CuentaPorCobrar.documento_id)
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(CuentaPorCobrar.pagado == False)
        .count()
    )

    clientes_activos = db.query(Cliente).filter(
        Cliente.empresa_id == empresa_id, Cliente.activo == True
    ).count()

    return {
        "productos_stock": productos_stock,
        "ventas_hoy": float(ventas_hoy or 0),
        "documentos_hoy": docs_hoy,
        "cartera_total": float(cartera or 0),
        "documentos_pendientes": docs_pendientes,
        "clientes_activos": clientes_activos,
    }


@router.get("/corte-caja")
def corte_caja(
    fecha: date | None = None,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    f = fecha or datetime.utcnow().date()
    inicio = datetime.combine(f, datetime.min.time())
    fin = inicio + timedelta(days=1)
    rows = (
        db.query(
            DocumentoVenta.tipo,
            func.count().label("n"),
            func.sum(DocumentoVenta.total).label("total"),
        )
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(DocumentoVenta.fecha >= inicio, DocumentoVenta.fecha < fin)
        .filter(DocumentoVenta.estatus != EstatusDocumento.CANCELADO.value)
        .group_by(DocumentoVenta.tipo)
        .all()
    )
    return {
        "fecha": f.isoformat(),
        "por_tipo": [
            {"tipo": t, "n": n, "total": float(total or 0)}
            for t, n, total in rows
        ],
    }


def _corte_data(db: Session, empresa_id: int, ini: datetime, fin: datetime) -> dict:
    """Calcula desglose de pagos del periodo dado, sin guardar."""
    # Ventas no canceladas en el rango
    ventas = (
        db.query(DocumentoVenta)
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(DocumentoVenta.fecha >= ini, DocumentoVenta.fecha < fin)
        .filter(DocumentoVenta.estatus != EstatusDocumento.CANCELADO.value)
        .filter(DocumentoVenta.tipo.in_([
            TipoDocumento.TICKET.value, TipoDocumento.FACTURA.value,
        ]))
        .all()
    )
    n_ventas = len(ventas)
    total_vendido = sum(float(v.total) for v in ventas)

    # Pagos: cuando se mando split, esta el detalle; si no, asumir 1 fila = total con la forma del doc
    desglose: dict[str, dict] = {}
    venta_ids = [v.id for v in ventas]
    pagos_rows = []
    if venta_ids:
        pagos_rows = db.query(Pago).filter(Pago.documento_venta_id.in_(venta_ids)).all()

    pagos_por_venta: dict[int, list[Pago]] = {}
    for p in pagos_rows:
        pagos_por_venta.setdefault(p.documento_venta_id, []).append(p)

    for v in ventas:
        rows = pagos_por_venta.get(v.id) or []
        if rows:
            for p in rows:
                k = p.forma_pago_sat
                d = desglose.setdefault(k, {"label": FORMA_LABEL.get(k, k), "monto": 0.0, "n": 0})
                d["monto"] += float(p.monto)
                d["n"] += 1
        else:
            # No hay split, contar el total con la forma del documento
            k = v.forma_pago_sat or "01"
            d = desglose.setdefault(k, {"label": FORMA_LABEL.get(k, k), "monto": 0.0, "n": 0})
            d["monto"] += float(v.total)
            d["n"] += 1

    efectivo_esperado = desglose.get("01", {}).get("monto", 0.0)
    return {
        "n_ventas": n_ventas,
        "total_vendido": round(total_vendido, 2),
        "desglose_pagos": desglose,
        "efectivo_esperado": round(efectivo_esperado, 2),
    }


@router.get("/corte/preview")
def corte_preview(
    fecha: date | None = None,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """Resumen del dia (sin guardar) - lo que veria el cajero antes de cerrar."""
    f = fecha or datetime.utcnow().date()
    ini = datetime.combine(f, datetime.min.time())
    fin = ini + timedelta(days=1)
    data = _corte_data(db, empresa_id, ini, fin)
    return {"fecha": f.isoformat(), **data}


class CerrarCorteIn(BaseModel):
    fecha: date | None = None
    efectivo_real: float
    notas: str | None = None


@router.post("/corte/cerrar")
def corte_cerrar(
    payload: CerrarCorteIn,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_current_user),
):
    """Cierra el corte: guarda snapshot + diferencia."""
    f = payload.fecha or datetime.utcnow().date()
    ini = datetime.combine(f, datetime.min.time())
    fin = ini + timedelta(days=1)
    data = _corte_data(db, empresa_id, ini, fin)

    efectivo_esp = data["efectivo_esperado"]
    diferencia = round(payload.efectivo_real - efectivo_esp, 2)

    corte = CorteCaja(
        empresa_id=empresa_id,
        usuario_id=usuario.id,
        fecha_corte=datetime.utcnow(),
        fecha_desde=ini,
        fecha_hasta=fin,
        n_ventas=data["n_ventas"],
        total_vendido=data["total_vendido"],
        desglose_pagos=data["desglose_pagos"],
        efectivo_esperado=efectivo_esp,
        efectivo_real=payload.efectivo_real,
        diferencia=diferencia,
        notas=payload.notas,
    )
    db.add(corte)
    db.commit()
    db.refresh(corte)
    return {
        "id": corte.id,
        "fecha": f.isoformat(),
        "diferencia": diferencia,
        "total_vendido": data["total_vendido"],
    }


@router.get("/corte/historial")
def corte_historial(
    limit: int = 60,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(CorteCaja)
        .filter(CorteCaja.empresa_id == empresa_id)
        .order_by(CorteCaja.fecha_corte.desc())
        .limit(limit)
        .all()
    )
    out = []
    for c in rows:
        u = db.get(Usuario, c.usuario_id) if c.usuario_id else None
        out.append({
            "id": c.id,
            "fecha": c.fecha_desde.date().isoformat(),
            "fecha_corte": c.fecha_corte.isoformat(),
            "usuario": u.nombre if u else "—",
            "n_ventas": c.n_ventas,
            "total_vendido": float(c.total_vendido),
            "efectivo_esperado": float(c.efectivo_esperado),
            "efectivo_real": float(c.efectivo_real),
            "diferencia": float(c.diferencia),
            "desglose_pagos": c.desglose_pagos,
            "notas": c.notas,
        })
    return out


# ===== Exportes XLSX =====

@router.get("/ventas-xlsx")
def export_ventas_xlsx(
    desde: date | None = None,
    hasta: date | None = None,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    hoy = datetime.utcnow().date()
    ini = datetime.combine(desde or hoy.replace(day=1), datetime.min.time())
    fin = datetime.combine(hasta or hoy, datetime.min.time()) + timedelta(days=1)

    q = (
        db.query(DocumentoVenta)
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(DocumentoVenta.fecha >= ini, DocumentoVenta.fecha < fin)
        .order_by(DocumentoVenta.fecha)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    headers = ["Folio", "Tipo", "Fecha", "Cliente", "RFC cliente",
               "Subtotal", "IVA", "Total", "Forma pago", "Método pago",
               "Estatus", "Notas"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")

    for d in q:
        cli = db.get(Cliente, d.cliente_id)
        ws.append([
            d.folio, d.tipo, d.fecha.strftime("%Y-%m-%d %H:%M"),
            cli.nombre if cli else "",
            (cli.rfc if cli else "") or "",
            float(d.subtotal), float(d.iva), float(d.total),
            FORMA_LABEL.get(d.forma_pago_sat, d.forma_pago_sat),
            d.metodo_pago_sat,
            d.estatus, d.notas or "",
        ])

    # Anchos de columna
    for col, w in zip("ABCDEFGHIJKL", [14, 12, 18, 32, 14, 12, 12, 14, 16, 8, 12, 30]):
        ws.column_dimensions[col].width = w

    # Resumen abajo
    if q:
        total_row = ws.max_row + 2
        ws.cell(row=total_row, column=7, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=sum(float(d.total) for d in q if d.estatus != "CANCELADO")).font = Font(bold=True)

    nombre = f"ventas_{ini.date().isoformat()}_{(fin - timedelta(days=1)).date().isoformat()}.xlsx"
    return _xlsx_response(wb, nombre)


@router.get("/cartera-xlsx")
def export_cartera_xlsx(
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = (
        db.query(CuentaPorCobrar)
        .join(DocumentoVenta, DocumentoVenta.id == CuentaPorCobrar.documento_id)
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(CuentaPorCobrar.pagado == False)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera"
    headers = ["Folio doc", "Cliente", "RFC", "Fecha emisión",
               "Vencimiento", "Días", "Monto original", "Saldo"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")

    today = datetime.utcnow().date()
    for cxc in rows:
        doc = db.get(DocumentoVenta, cxc.documento_id)
        cli = db.get(Cliente, cxc.cliente_id)
        dias = (today - cxc.fecha_emision.date()).days
        venc = cxc.fecha_vencimiento.date().isoformat() if cxc.fecha_vencimiento else ""
        ws.append([
            doc.folio if doc else "",
            cli.nombre if cli else "",
            (cli.rfc if cli else "") or "",
            cxc.fecha_emision.date().isoformat(),
            venc, dias,
            float(cxc.monto_original), float(cxc.saldo),
        ])

    for col, w in zip("ABCDEFGH", [16, 32, 14, 14, 14, 8, 14, 14]):
        ws.column_dimensions[col].width = w

    if rows:
        total_row = ws.max_row + 2
        ws.cell(row=total_row, column=7, value="TOTAL SALDO").font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=sum(float(c.saldo) for c in rows)).font = Font(bold=True)

    return _xlsx_response(wb, f"cartera_{today.isoformat()}.xlsx")


@router.get("/cortes-xlsx")
def export_cortes_xlsx(
    desde: date | None = None,
    hasta: date | None = None,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    hoy = datetime.utcnow().date()
    ini = datetime.combine(desde or hoy.replace(day=1), datetime.min.time())
    fin = datetime.combine(hasta or hoy, datetime.min.time()) + timedelta(days=1)

    rows = (
        db.query(CorteCaja)
        .filter(CorteCaja.empresa_id == empresa_id)
        .filter(CorteCaja.fecha_desde >= ini, CorteCaja.fecha_desde < fin)
        .order_by(CorteCaja.fecha_desde)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cortes"
    headers = ["Fecha", "Cajero", "# Ventas", "Total vendido",
               "Efectivo", "Transferencia", "T. Crédito", "T. Débito",
               "Efectivo esperado", "Efectivo real", "Diferencia", "Notas"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")

    for c in rows:
        u = db.get(Usuario, c.usuario_id) if c.usuario_id else None
        d = c.desglose_pagos or {}
        ws.append([
            c.fecha_desde.date().isoformat(),
            u.nombre if u else "",
            c.n_ventas, float(c.total_vendido),
            d.get("01", {}).get("monto", 0),
            d.get("03", {}).get("monto", 0),
            d.get("04", {}).get("monto", 0),
            d.get("28", {}).get("monto", 0),
            float(c.efectivo_esperado), float(c.efectivo_real), float(c.diferencia),
            c.notas or "",
        ])

    for col, w in zip("ABCDEFGHIJKL", [12, 22, 10, 14, 12, 14, 12, 12, 16, 14, 12, 28]):
        ws.column_dimensions[col].width = w

    return _xlsx_response(wb, f"cortes_{ini.date().isoformat()}_{(fin - timedelta(days=1)).date().isoformat()}.xlsx")


@router.get("/antiguedad-cartera")
def antiguedad_cartera(
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    today = datetime.utcnow().date()
    buckets = {"0-15": 0.0, "16-30": 0.0, "31-60": 0.0, "61-90": 0.0, "91+": 0.0}
    rows = (
        db.query(CuentaPorCobrar)
        .join(DocumentoVenta, DocumentoVenta.id == CuentaPorCobrar.documento_id)
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(CuentaPorCobrar.pagado == False)
        .all()
    )
    for cxc in rows:
        d = (today - cxc.fecha_emision.date()).days
        saldo = float(cxc.saldo)
        if d <= 15: buckets["0-15"] += saldo
        elif d <= 30: buckets["16-30"] += saldo
        elif d <= 60: buckets["31-60"] += saldo
        elif d <= 90: buckets["61-90"] += saldo
        else: buckets["91+"] += saldo
    return buckets


# === Reporte diario de facturas agrupado por forma de pago ===

def _reporte_diario_data(db: Session, empresa_id: int, ini: datetime, fin: datetime, incluir_tickets: bool = False):
    """Construye los grupos por forma de pago: facturas del rango."""
    from app.models import Cfdi
    tipos = [TipoDocumento.FACTURA.value]
    if incluir_tickets:
        tipos.append(TipoDocumento.TICKET.value)
    rows = (
        db.query(DocumentoVenta, Cliente)
        .join(Cliente, Cliente.id == DocumentoVenta.cliente_id)
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(DocumentoVenta.fecha >= ini, DocumentoVenta.fecha < fin)
        .filter(DocumentoVenta.estatus != EstatusDocumento.CANCELADO.value)
        .filter(DocumentoVenta.tipo.in_(tipos))
        .order_by(DocumentoVenta.forma_pago_sat, DocumentoVenta.fecha)
        .all()
    )

    # Agrupar por forma_pago_sat
    grupos: dict[str, list[dict]] = {}
    for d, cli in rows:
        # Para FACTURA, tomar serie/folio del CFDI si esta timbrada
        cfdi = db.query(Cfdi).filter(
            Cfdi.documento_venta_id == d.id, Cfdi.cancelado == False
        ).first()
        # Estado: CO=cobrado (PUE timbrada), PE=pendiente (PPD sin pagar), CR=cancelado
        if d.metodo_pago_sat == "PPD":
            cxc = db.query(CuentaPorCobrar).filter(
                CuentaPorCobrar.documento_id == d.id, CuentaPorCobrar.pagado == False
            ).first()
            estado = "PE" if cxc else "CO"
        else:
            estado = "CO"
        forma = d.forma_pago_sat or "01"
        if forma not in grupos:
            grupos[forma] = []
        grupos[forma].append({
            "fecha": d.fecha.strftime("%d/%m/%Y"),
            "serie": cfdi.serie if cfdi else (("FC" if d.tipo == "FACTURA" else "TK")),
            "referencia": cfdi.folio if cfdi else d.folio,
            "folio_interno": d.folio,
            "cliente": (cli.razon_social or cli.nombre).upper(),
            "rfc": cli.rfc or "",
            "total": float(d.total),
            "estado": estado,
            "metodo": d.metodo_pago_sat,
        })

    # Ordenar grupos por codigo SAT
    return dict(sorted(grupos.items(), key=lambda kv: kv[0])), len(rows), sum(float(d.total) for d, _ in rows)


@router.get("/diario")
def reporte_diario(
    desde: date | None = None,
    hasta: date | None = None,
    incluir_tickets: bool = False,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    f_desde = desde or datetime.utcnow().date()
    f_hasta = hasta or f_desde
    ini = datetime.combine(f_desde, datetime.min.time())
    fin = datetime.combine(f_hasta, datetime.min.time()) + timedelta(days=1)

    grupos, n_total, total_gral = _reporte_diario_data(db, empresa_id, ini, fin, incluir_tickets)

    return {
        "desde": f_desde.isoformat(),
        "hasta": f_hasta.isoformat(),
        "grupos": [
            {
                "forma_pago": k,
                "label": FORMA_LABEL.get(k, k),
                "n": len(rows),
                "subtotal": round(sum(r["total"] for r in rows), 2),
                "facturas": rows,
            }
            for k, rows in grupos.items()
        ],
        "n_total": n_total,
        "total_general": round(total_gral, 2),
    }


@router.get("/diario-xlsx")
def reporte_diario_xlsx(
    desde: date | None = None,
    hasta: date | None = None,
    incluir_tickets: bool = False,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    f_desde = desde or datetime.utcnow().date()
    f_hasta = hasta or f_desde
    ini = datetime.combine(f_desde, datetime.min.time())
    fin = datetime.combine(f_hasta, datetime.min.time()) + timedelta(days=1)
    grupos, n_total, total_gral = _reporte_diario_data(db, empresa_id, ini, fin, incluir_tickets)

    # Empresa
    from app.models import Empresa
    emp = db.get(Empresa, empresa_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Header titulo
    ws["A1"] = (emp.nombre if emp else "ACEROMAX").upper()
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")

    rango = f"Del día {f_desde.strftime('%d-%m-%Y')} al día {f_hasta.strftime('%d-%m-%Y')}"
    ws["A2"] = f"Reporte de Facturas — {rango}"
    ws["A2"].font = Font(bold=True, size=11)
    ws.merge_cells("A2:F2")

    # Encabezado columnas
    headers = ["Fecha", "Serie", "Referencia", "Cliente", "Total", "Estado"]
    row_idx = 4
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.alignment = Alignment(horizontal="left")
    thin = Side(style="thin", color="9CA3AF")

    row_idx = 5
    for forma_sat, rows in grupos.items():
        # Label de grupo
        label_grupo = f"Pago {forma_sat}" if forma_sat in ("01", "02", "03") else (
            "Pago TA" if forma_sat in ("04", "28") else f"Pago {forma_sat}"
        )
        ws.cell(row=row_idx, column=1, value=label_grupo).font = Font(bold=True, italic=True)
        row_idx += 1
        for r in rows:
            ws.cell(row=row_idx, column=1, value=r["fecha"])
            ws.cell(row=row_idx, column=2, value=r["serie"])
            ws.cell(row=row_idx, column=3, value=r["referencia"])
            ws.cell(row=row_idx, column=4, value=r["cliente"])
            ws.cell(row=row_idx, column=5, value=r["total"]).number_format = '"$"#,##0.00'
            ws.cell(row=row_idx, column=6, value=r["estado"])
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = Border(top=thin, bottom=thin)
            row_idx += 1
        # Subtotal del grupo
        ws.cell(row=row_idx, column=1, value=f"{len(rows)} doc(s)").font = Font(bold=True)
        ws.cell(row=row_idx, column=5, value=round(sum(r["total"] for r in rows), 2)).number_format = '"$"#,##0.00'
        ws.cell(row=row_idx, column=5).font = Font(bold=True)
        ws.cell(row=row_idx, column=5).fill = PatternFill("solid", fgColor="F3F4F6")
        row_idx += 2

    # Total general
    ws.cell(row=row_idx, column=1, value=f"TOTAL: {n_total} doc(s)").font = Font(bold=True, size=12)
    ws.cell(row=row_idx, column=5, value=round(total_gral, 2)).number_format = '"$"#,##0.00'
    ws.cell(row=row_idx, column=5).font = Font(bold=True, size=12)
    ws.cell(row=row_idx, column=5).fill = PatternFill("solid", fgColor="DBEAFE")

    # Anchos
    for col, w in zip("ABCDEF", [12, 8, 14, 42, 14, 8]):
        ws.column_dimensions[col].width = w

    nombre = f"reporte_facturas_{f_desde.isoformat()}_{f_hasta.isoformat()}.xlsx"
    return _xlsx_response(wb, nombre)
