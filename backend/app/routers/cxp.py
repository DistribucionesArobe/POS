"""Compras y cuentas por pagar - filtrado por empresa."""
from datetime import datetime, date, timedelta
try:
    from zoneinfo import ZoneInfo
    TZ_MX = ZoneInfo("America/Mexico_City")
except Exception:
    TZ_MX = None


def _hoy_mx() -> date:
    """Fecha de hoy en zona horaria de Mexico (no UTC)."""
    if TZ_MX:
        return datetime.now(TZ_MX).date()
    return datetime.utcnow().date()


def _dias_habiles_restantes_semana(d: date) -> int:
    """Cuenta dias laborales (lun-vie) restantes desde d INCLUSIVE hasta el viernes.
    Mar=4, Vie=1, Sab/Dom=5 (toma la siguiente semana completa).
    """
    wd = d.weekday()  # 0=lun, 4=vie, 5=sab, 6=dom
    if wd <= 4:
        return 5 - wd
    return 5  # fin de semana: cuenta lun-vie de la proxima semana
from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, Field
from sqlalchemy import or_, func
from sqlalchemy.orm import Session, joinedload

from app.db import get_db
from app.models import (
    CuentaPorPagar, Compra, Proveedor, PanelCxP, AbonoCxP,
    DocumentoVenta, CuentaPorCobrar, OtroPagoPanel,
)
from app.models.venta import EstatusDocumento, TipoDocumento
from app.schemas.compra import CompraIn, AbonoCxPIn
from app.services import compra_service
from app.services.security import get_active_empresa_id

router = APIRouter()


@router.get("/cartera")
def cartera_proveedores(
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
    incluir_pagadas: bool = False,
    anio: int | None = None,
    mes: int | None = None,
    arrastrar_vencidas: bool = True,  # CxP vencidas de meses anteriores tambien aparecen
):
    """Lista CxP. Incluye tanto las ligadas a una Compra como las manuales.

    Comportamiento del filtro por mes:
    - CxP que vencen EN ese mes (default)
    - + CxP vencidas en meses anteriores que aun no se pagan (arrastrar_vencidas=True)
    - + CxP sin fecha de vencimiento que llegaron en ese mes

    Esto replica el flujo Excel de 'copy/paste de pendientes al mes siguiente'
    pero automatico: las facturas no pagadas te siguen mes a mes hasta saldarlas.
    """
    q = (
        db.query(CuentaPorPagar, Proveedor)
        .join(Proveedor, Proveedor.id == CuentaPorPagar.proveedor_id)
        .outerjoin(Compra, Compra.id == CuentaPorPagar.compra_id)
        .filter(or_(
            CuentaPorPagar.empresa_id == empresa_id,
            Compra.empresa_id == empresa_id,
        ))
    )
    if not incluir_pagadas:
        q = q.filter(CuentaPorPagar.pagado == False)
    if anio and mes:
        from sqlalchemy import and_
        ini = datetime(anio, mes, 1)
        fin = datetime(anio + 1, 1, 1) if mes == 12 else datetime(anio, mes + 1, 1)
        filtros_mes = [
            # Vencen EN el mes
            CuentaPorPagar.fecha_vencimiento.between(ini, fin),
            # Sin fecha de vencimiento pero recibida en el mes
            and_(
                CuentaPorPagar.fecha_vencimiento.is_(None),
                CuentaPorPagar.fecha_recepcion.between(ini, fin),
            ),
        ]
        if arrastrar_vencidas:
            # CxP vencidas antes del mes y aun no pagadas (se arrastran)
            filtros_mes.append(and_(
                CuentaPorPagar.fecha_vencimiento < ini,
                CuentaPorPagar.pagado == False,
            ))
        q = q.filter(or_(*filtros_mes))

    rows = q.all()
    out = []
    for cxp, prov in rows:
        compra = db.get(Compra, cxp.compra_id) if cxp.compra_id else None
        out.append({
            "cxp_id": cxp.id, "proveedor_id": prov.id, "proveedor": prov.nombre,
            "compra_id": cxp.compra_id,
            "compra_folio": compra.folio_interno if compra else None,
            "folio_factura": cxp.folio_factura,
            "fecha_recepcion": cxp.fecha_recepcion.isoformat() if cxp.fecha_recepcion else None,
            "observaciones": cxp.observaciones,
            "moneda": getattr(cxp, "moneda", "MXN") or "MXN",
            "tipo_cambio": float(cxp.tipo_cambio) if getattr(cxp, "tipo_cambio", None) else None,
            "monto_moneda_original": float(cxp.monto_moneda_original) if getattr(cxp, "monto_moneda_original", None) else None,
            "corto_plazo": bool(getattr(cxp, "corto_plazo", False)),
            "monto_original": float(cxp.monto_original),
            "saldo": float(cxp.saldo),
            "saldado": float(cxp.monto_original) - float(cxp.saldo),
            "fecha_vencimiento": cxp.fecha_vencimiento.isoformat() if cxp.fecha_vencimiento else None,
            "pagado": cxp.pagado,
            "manual": cxp.compra_id is None,
        })
    # Orden por fecha vencimiento
    out.sort(key=lambda x: x.get("fecha_vencimiento") or x.get("fecha_recepcion") or "")
    return out


# ===== CxP manual (estilo Excel) =====

class CxPManualIn(BaseModel):
    proveedor_id: int | None = None
    proveedor_nombre: str | None = None  # si no existe, lo creamos
    folio_factura: str | None = None
    fecha_recepcion: date | None = None
    fecha_vencimiento: date | None = None
    monto_original: float = Field(gt=0)
    saldado: float = Field(default=0, ge=0)
    observaciones: str | None = None
    # Moneda: MXN o USD. Si USD, se requiere tipo_cambio.
    moneda: str = "MXN"
    tipo_cambio: float | None = None


@router.post("/manual")
def crear_cxp_manual(
    payload: CxPManualIn,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """Crea una CxP sin compra ligada (captura rapida estilo Excel)."""
    # Resolver proveedor
    if payload.proveedor_id:
        prov = db.get(Proveedor, payload.proveedor_id)
        if not prov or prov.empresa_id != empresa_id:
            raise HTTPException(400, "Proveedor no existe o de otra empresa")
    elif payload.proveedor_nombre:
        # Buscar por nombre, si no existe crear
        prov = (
            db.query(Proveedor)
            .filter(Proveedor.empresa_id == empresa_id)
            .filter(Proveedor.nombre.ilike(payload.proveedor_nombre.strip()))
            .first()
        )
        if not prov:
            prov = Proveedor(
                empresa_id=empresa_id,
                nombre=payload.proveedor_nombre.strip(),
            )
            db.add(prov)
            db.flush()
    else:
        raise HTTPException(400, "Captura proveedor_id o proveedor_nombre")

    # Procesar moneda + tipo de cambio
    moneda = (payload.moneda or "MXN").upper()
    if moneda not in ("MXN", "USD"):
        raise HTTPException(400, "moneda debe ser MXN o USD")
    if moneda == "USD" and (not payload.tipo_cambio or payload.tipo_cambio <= 0):
        raise HTTPException(400, "Para moneda USD se requiere tipo_cambio")

    # Conversion a MXN para los totales del tablero
    if moneda == "USD":
        monto_mxn = round(payload.monto_original * payload.tipo_cambio, 2)
        saldado_mxn = round(payload.saldado * payload.tipo_cambio, 2)
        monto_moneda_original = payload.monto_original  # en USD
    else:
        monto_mxn = payload.monto_original
        saldado_mxn = payload.saldado
        monto_moneda_original = None

    if saldado_mxn > monto_mxn:
        raise HTTPException(400, "Saldado no puede exceder el monto original")

    saldo = round(monto_mxn - saldado_mxn, 2)
    # Fecha de recepcion: la del payload o hoy
    fecha_recep = (
        datetime.combine(payload.fecha_recepcion, datetime.min.time())
        if payload.fecha_recepcion else datetime.utcnow()
    )
    # Vencimiento: si no se mando, default a +30 dias de la recepcion (editable)
    if payload.fecha_vencimiento:
        fecha_venc = datetime.combine(payload.fecha_vencimiento, datetime.min.time())
    else:
        fecha_venc = fecha_recep + timedelta(days=30)
    cxp = CuentaPorPagar(
        empresa_id=empresa_id,
        proveedor_id=prov.id,
        compra_id=None,
        folio_factura=payload.folio_factura,
        fecha_recepcion=fecha_recep,
        fecha_vencimiento=fecha_venc,
        observaciones=payload.observaciones,
        moneda=moneda,
        tipo_cambio=payload.tipo_cambio if moneda == "USD" else None,
        monto_moneda_original=monto_moneda_original,
        monto_original=monto_mxn,
        saldo=saldo,
        pagado=saldo <= 0.01,
    )
    db.add(cxp)
    db.commit()
    db.refresh(cxp)
    return {"id": cxp.id, "proveedor": prov.nombre, "saldo": float(cxp.saldo)}


@router.patch("/manual/{cxp_id}")
def actualizar_cxp_manual(
    cxp_id: int,
    payload: dict,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """Editar campos de una CxP (folio_factura, observaciones, fechas, saldado)."""
    cxp = db.get(CuentaPorPagar, cxp_id)
    if not cxp:
        raise HTTPException(404, "CxP no existe")
    # Validar empresa (via compra o empresa_id directo)
    if cxp.empresa_id and cxp.empresa_id != empresa_id:
        raise HTTPException(403, "CxP de otra empresa")
    if cxp.compra_id:
        compra = db.get(Compra, cxp.compra_id)
        if compra and compra.empresa_id != empresa_id:
            raise HTTPException(403, "CxP de otra empresa")

    if "folio_factura" in payload:
        cxp.folio_factura = payload["folio_factura"] or None
    if "observaciones" in payload:
        cxp.observaciones = payload["observaciones"] or None
    # Detectar si vencimiento explicitamente viene en el payload (aunque sea null)
    venc_in_payload = "fecha_vencimiento" in payload
    recep_in_payload = "fecha_recepcion" in payload

    if venc_in_payload and payload["fecha_vencimiento"]:
        cxp.fecha_vencimiento = datetime.fromisoformat(payload["fecha_vencimiento"])
    if recep_in_payload and payload["fecha_recepcion"]:
        nueva_recep = datetime.fromisoformat(payload["fecha_recepcion"])
        cxp.fecha_recepcion = nueva_recep
        # Si el usuario cambio la recepcion pero NO mando vencimiento en este PATCH,
        # auto-mover vencimiento a recepcion + 30 dias (puede ser nuevo o reajuste).
        if not venc_in_payload:
            cxp.fecha_vencimiento = nueva_recep + timedelta(days=30)
    if "saldado" in payload:
        saldado = float(payload["saldado"])
        if saldado > float(cxp.monto_original):
            raise HTTPException(400, "Saldado no puede exceder el monto original")
        cxp.saldo = round(float(cxp.monto_original) - saldado, 2)
        cxp.pagado = cxp.saldo <= 0.01
    if "monto_original" in payload:
        try:
            nuevo_monto = float(payload["monto_original"])
        except (TypeError, ValueError):
            raise HTTPException(400, "Monto invalido")
        if nuevo_monto <= 0:
            raise HTTPException(400, "Monto debe ser mayor a 0")
        # Conserva lo ya saldado y recalcula el saldo pendiente
        saldado_actual = float(cxp.monto_original) - float(cxp.saldo)
        if saldado_actual > nuevo_monto:
            raise HTTPException(
                400,
                f"No puedes bajar el monto a {nuevo_monto:.2f} porque ya tiene saldado {saldado_actual:.2f}. "
                "Reversa los abonos primero o usa un monto mayor."
            )
        cxp.monto_original = nuevo_monto
        cxp.saldo = round(nuevo_monto - saldado_actual, 2)
        cxp.pagado = cxp.saldo <= 0.01
    if "corto_plazo" in payload:
        cxp.corto_plazo = bool(payload["corto_plazo"])
    db.commit()
    return {"ok": True}


@router.delete("/manual/{cxp_id}")
def borrar_cxp_manual(
    cxp_id: int,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """Solo se pueden borrar CxP manuales (las ligadas a Compra hay que cancelar la compra)."""
    cxp = db.get(CuentaPorPagar, cxp_id)
    if not cxp:
        raise HTTPException(404, "CxP no existe")
    if cxp.compra_id is not None:
        raise HTTPException(400, "Esta CxP esta ligada a una Compra; no se puede borrar suelta")
    if cxp.empresa_id and cxp.empresa_id != empresa_id:
        raise HTTPException(403, "CxP de otra empresa")
    db.delete(cxp)
    db.commit()
    return {"ok": True}


# ===== Panel mensual (tablero estilo Excel) =====

class PanelIn(BaseModel):
    anio: int
    mes: int
    venta_objetivo_mes: float = 0
    saldo_banco: float = 0
    ingreso_egreso_banco: float = 0
    usd_mxn: float = 0
    notas: str | None = None
    ingreso_mensual: float = 0
    errores_mensual: float = 0


@router.get("/tablero")
def tablero_cxp(
    anio: int = Query(...),
    mes: int = Query(..., ge=1, le=12),
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """Devuelve panel + KPIs calculados + lista de CxP del mes."""
    # Panel (upsert pasivo - lo trae aunque no exista)
    panel = (
        db.query(PanelCxP)
        .filter(PanelCxP.empresa_id == empresa_id)
        .filter(PanelCxP.anio == anio, PanelCxP.mes == mes)
        .first()
    )
    if not panel:
        panel_dict = {
            "anio": anio, "mes": mes,
            "venta_objetivo_mes": 0, "saldo_banco": 0,
            "ingreso_egreso_banco": 0, "usd_mxn": 0, "notas": None,
            "ingreso_mensual": 0, "errores_mensual": 0,
        }
    else:
        panel_dict = {
            "id": panel.id, "anio": panel.anio, "mes": panel.mes,
            "venta_objetivo_mes": float(panel.venta_objetivo_mes),
            "saldo_banco": float(panel.saldo_banco),
            "ingreso_egreso_banco": float(getattr(panel, "ingreso_egreso_banco", 0) or 0),
            "usd_mxn": float(panel.usd_mxn),
            "notas": panel.notas,
            "ingreso_mensual": float(getattr(panel, "ingreso_mensual", 0) or 0),
            "errores_mensual": float(getattr(panel, "errores_mensual", 0) or 0),
            "actualizado_en": panel.actualizado_en.isoformat(),
        }

    # KPIs auto-calculados del POS
    ini_mes = datetime(anio, mes, 1)
    fin_mes = datetime(anio + 1, 1, 1) if mes == 12 else datetime(anio, mes + 1, 1)
    hoy = _hoy_mx()  # zona horaria Mexico
    dias_mes = (fin_mes - ini_mes).days
    if ini_mes.date() <= hoy < fin_mes.date():
        dia_actual = (hoy - ini_mes.date()).days + 1
    else:
        dia_actual = dias_mes

    # Ventas del mes - INGRESO bruto (ticket/factura/remision)
    venta_ingreso = (
        db.query(func.coalesce(func.sum(DocumentoVenta.total), 0))
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(DocumentoVenta.fecha >= ini_mes, DocumentoVenta.fecha < fin_mes)
        .filter(DocumentoVenta.tipo.in_([
            TipoDocumento.TICKET.value, TipoDocumento.FACTURA.value, TipoDocumento.REMISION.value,
        ]))
        .filter(DocumentoVenta.estatus != EstatusDocumento.CANCELADO.value)
        .scalar()
    )
    venta_ingreso = float(venta_ingreso or 0)

    # DEVOLUCIONES del mes (notas de credito) - se restan del ingreso
    venta_devoluciones = (
        db.query(func.coalesce(func.sum(DocumentoVenta.total), 0))
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(DocumentoVenta.fecha >= ini_mes, DocumentoVenta.fecha < fin_mes)
        .filter(DocumentoVenta.tipo == "NOTA_CREDITO")
        .filter(DocumentoVenta.estatus != EstatusDocumento.CANCELADO.value)
        .scalar()
    )
    venta_devoluciones = float(venta_devoluciones or 0)

    # Venta NETA del mes = Ingreso - Devoluciones
    venta_mes = venta_ingreso - venta_devoluciones

    # Total CxP abiertas (todas, no solo del mes)
    cxp_total = (
        db.query(func.coalesce(func.sum(CuentaPorPagar.saldo), 0))
        .outerjoin(Compra, Compra.id == CuentaPorPagar.compra_id)
        .filter(or_(
            CuentaPorPagar.empresa_id == empresa_id,
            Compra.empresa_id == empresa_id,
        ))
        .filter(CuentaPorPagar.pagado == False)
        .scalar()
    )
    cxp_total = float(cxp_total or 0)

    # CxP vencen este mes
    cxp_del_mes = (
        db.query(func.coalesce(func.sum(CuentaPorPagar.saldo), 0))
        .outerjoin(Compra, Compra.id == CuentaPorPagar.compra_id)
        .filter(or_(
            CuentaPorPagar.empresa_id == empresa_id,
            Compra.empresa_id == empresa_id,
        ))
        .filter(CuentaPorPagar.pagado == False)
        .filter(CuentaPorPagar.fecha_vencimiento.between(ini_mes, fin_mes))
        .scalar()
    )
    cxp_del_mes = float(cxp_del_mes or 0)

    # CxP marcadas como corto plazo (las que el usuario va a pagar pronto)
    cxp_corto_plazo = (
        db.query(func.coalesce(func.sum(CuentaPorPagar.saldo), 0))
        .outerjoin(Compra, Compra.id == CuentaPorPagar.compra_id)
        .filter(or_(
            CuentaPorPagar.empresa_id == empresa_id,
            Compra.empresa_id == empresa_id,
        ))
        .filter(CuentaPorPagar.pagado == False)
        .filter(CuentaPorPagar.corto_plazo == True)
        .scalar()
    )
    cxp_corto_plazo = float(cxp_corto_plazo or 0)

    # Total CxC abierta
    cxc_total = (
        db.query(func.coalesce(func.sum(CuentaPorCobrar.saldo), 0))
        .join(DocumentoVenta, DocumentoVenta.id == CuentaPorCobrar.documento_id)
        .filter(DocumentoVenta.empresa_id == empresa_id)
        .filter(CuentaPorCobrar.pagado == False)
        .scalar()
    )
    cxc_total = float(cxc_total or 0)

    # Otros pagos del tablero (renta, sueldos, servicios, etc.)
    otros_pagos_total = (
        db.query(func.coalesce(func.sum(OtroPagoPanel.monto), 0))
        .filter(OtroPagoPanel.empresa_id == empresa_id)
        .scalar()
    )
    otros_pagos_total = float(otros_pagos_total or 0)

    # Cálculos del Excel
    venta_objetivo = panel_dict["venta_objetivo_mes"]
    restante_meta = max(0, venta_objetivo - venta_mes)
    dias_restantes = max(1, dias_mes - dia_actual + 1)
    venta_promedio_dia = (venta_mes / dia_actual) if dia_actual else 0
    venta_estimada_mes = venta_promedio_dia * dias_mes
    a_vender_por_dia = restante_meta / dias_restantes
    diferencia = (cxc_total + venta_estimada_mes) - cxp_total

    # Dias habiles restantes hasta el viernes (para "a vender por dia")
    dias_habiles_semana = _dias_habiles_restantes_semana(hoy)

    return {
        "panel": panel_dict,
        "kpis": {
            "dia_actual": dia_actual,
            "dias_mes": dias_mes,
            "dias_restantes": dias_restantes,
            "dias_habiles_semana": dias_habiles_semana,
            "venta_mes": round(venta_mes, 2),  # NETA (ingreso - devoluciones)
            "venta_ingreso": round(venta_ingreso, 2),
            "venta_devoluciones": round(venta_devoluciones, 2),
            "venta_promedio_dia": round(venta_promedio_dia, 2),
            "venta_estimada_mes": round(venta_estimada_mes, 2),
            "restante_meta": round(restante_meta, 2),
            "a_vender_por_dia": round(a_vender_por_dia, 2),
            "cxp_total": round(cxp_total, 2),
            "cxp_del_mes": round(cxp_del_mes, 2),
            "cxp_corto_plazo": round(cxp_corto_plazo, 2),
            "cxc_total": round(cxc_total, 2),
            "diferencia": round(diferencia, 2),
            "otros_pagos_total": round(otros_pagos_total, 2),
        },
    }


# ===== Otros pagos del tablero (renta, sueldos, servicios, etc.) =====

@router.get("/otros-pagos")
def listar_otros_pagos(
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(OtroPagoPanel)
        .filter(OtroPagoPanel.empresa_id == empresa_id)
        .order_by(OtroPagoPanel.orden, OtroPagoPanel.id)
        .all()
    )
    return [
        {"id": r.id, "concepto": r.concepto, "monto": float(r.monto or 0), "orden": r.orden}
        for r in rows
    ]


@router.post("/otros-pagos")
def crear_otro_pago(
    payload: dict,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    concepto = (payload.get("concepto") or "").strip()
    if not concepto:
        raise HTTPException(400, "Concepto requerido")
    try:
        monto = float(payload.get("monto") or 0)
    except (TypeError, ValueError):
        raise HTTPException(400, "Monto invalido")
    siguiente = db.query(OtroPagoPanel).filter(OtroPagoPanel.empresa_id == empresa_id).count() + 1
    r = OtroPagoPanel(
        empresa_id=empresa_id, concepto=concepto, monto=monto, orden=siguiente,
    )
    db.add(r); db.commit(); db.refresh(r)
    return {"id": r.id, "concepto": r.concepto, "monto": float(r.monto), "orden": r.orden}


@router.patch("/otros-pagos/{pid}")
def actualizar_otro_pago(
    pid: int, payload: dict,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    r = db.get(OtroPagoPanel, pid)
    if not r or r.empresa_id != empresa_id:
        raise HTTPException(404, "Otro pago no existe")
    if "concepto" in payload:
        c = (payload["concepto"] or "").strip()
        if c:
            r.concepto = c
    if "monto" in payload:
        try:
            r.monto = float(payload["monto"])
        except (TypeError, ValueError):
            raise HTTPException(400, "Monto invalido")
    db.commit()
    return {"ok": True, "id": r.id}


@router.delete("/otros-pagos/{pid}")
def borrar_otro_pago(
    pid: int,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    r = db.get(OtroPagoPanel, pid)
    if not r or r.empresa_id != empresa_id:
        raise HTTPException(404, "Otro pago no existe")
    db.delete(r); db.commit()
    return {"ok": True}


@router.get("/deuda-por-proveedor")
def deuda_por_proveedor(
    incluir_zero: bool = True,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """Agrupa saldo pendiente por proveedor para el tablero estilo Excel.
    Si incluir_zero=true, devuelve TODOS los proveedores (incluso con saldo 0).
    """
    # Suma de saldo por proveedor
    sub = (
        db.query(
            CuentaPorPagar.proveedor_id,
            func.coalesce(func.sum(CuentaPorPagar.saldo), 0).label("saldo"),
        )
        .outerjoin(Compra, Compra.id == CuentaPorPagar.compra_id)
        .filter(or_(
            CuentaPorPagar.empresa_id == empresa_id,
            Compra.empresa_id == empresa_id,
        ))
        .filter(CuentaPorPagar.pagado == False)
        .group_by(CuentaPorPagar.proveedor_id)
        .subquery()
    )
    rows = (
        db.query(Proveedor, sub.c.saldo)
        .outerjoin(sub, sub.c.proveedor_id == Proveedor.id)
        .filter(Proveedor.empresa_id == empresa_id)
        .filter(Proveedor.activo == True)
        .all()
    )
    out = []
    total = 0.0
    for prov, saldo in rows:
        s = float(saldo or 0)
        if not incluir_zero and s <= 0.01:
            continue
        out.append({
            "proveedor_id": prov.id,
            "proveedor": prov.nombre,
            "saldo": s,
        })
        total += s
    out.sort(key=lambda x: -x["saldo"])
    for r in out:
        r["pct"] = (r["saldo"] / total * 100) if total > 0 else 0
    return {"total": round(total, 2), "filas": out}


@router.post("/panel")
def upsert_panel(
    payload: PanelIn,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """Crea o actualiza el panel mensual del usuario."""
    panel = (
        db.query(PanelCxP)
        .filter(PanelCxP.empresa_id == empresa_id)
        .filter(PanelCxP.anio == payload.anio, PanelCxP.mes == payload.mes)
        .first()
    )
    if not panel:
        panel = PanelCxP(empresa_id=empresa_id, anio=payload.anio, mes=payload.mes)
        db.add(panel)
    panel.venta_objetivo_mes = payload.venta_objetivo_mes
    panel.saldo_banco = payload.saldo_banco
    panel.ingreso_egreso_banco = payload.ingreso_egreso_banco
    panel.usd_mxn = payload.usd_mxn
    panel.notas = payload.notas
    panel.ingreso_mensual = payload.ingreso_mensual
    panel.errores_mensual = payload.errores_mensual
    panel.actualizado_en = datetime.utcnow()
    db.commit()
    return {"ok": True}


@router.get("/tipo-cambio")
def tipo_cambio_usd_mxn():
    """Obtiene tipo de cambio USD/MXN. Intenta Banxico SIE API (FIX oficial)
    si hay token, si no usa fallback gratis (open.er-api.com)."""
    import os, httpx

    banxico_token = os.environ.get("BANXICO_TOKEN", "").strip()

    if banxico_token:
        # Serie SF63528 = Tipo de cambio FIX
        url = "https://www.banxico.org.mx/SieAPIRest/service/v1/series/SF63528/datos/oportuno"
        try:
            r = httpx.get(url, headers={"Bmx-Token": banxico_token}, timeout=10)
            r.raise_for_status()
            data = r.json()
            serie = data["bmx"]["series"][0]
            dato = serie["datos"][0]
            return {
                "fuente": "Banxico FIX",
                "fecha": dato["fecha"],
                "valor": float(dato["dato"]),
            }
        except Exception as e:
            # cae al fallback
            pass

    # Fallback: open.er-api.com (gratis, sin token)
    try:
        r = httpx.get("https://open.er-api.com/v6/latest/USD", timeout=10)
        r.raise_for_status()
        data = r.json()
        mxn = data["rates"]["MXN"]
        ts = data.get("time_last_update_utc", "")
        return {
            "fuente": "open.er-api.com (no es FIX SAT)",
            "fecha": ts,
            "valor": float(mxn),
        }
    except Exception as e:
        raise HTTPException(503, f"No se pudo obtener tipo de cambio: {e}")


@router.get("/compras")
def listar_compras(
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(Compra)
        .filter(Compra.empresa_id == empresa_id)
        .order_by(Compra.fecha_recepcion.desc())
        .limit(100)
        .all()
    )
    return [
        {
            "id": c.id, "folio": c.folio_interno, "proveedor_id": c.proveedor_id,
            "uuid_cfdi": c.uuid_cfdi,
            "folio_factura_proveedor": c.folio_factura_proveedor,
            "subtotal": float(c.subtotal), "iva": float(c.iva), "total": float(c.total),
            "estatus": c.estatus,
            "fecha_recepcion": c.fecha_recepcion.isoformat(),
        }
        for c in rows
    ]


@router.get("/compras/{compra_id}")
def obtener_compra(
    compra_id: int,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    compra = (
        db.query(Compra)
        .options(joinedload(Compra.conceptos))
        .filter(Compra.id == compra_id, Compra.empresa_id == empresa_id)
        .first()
    )
    if not compra:
        raise HTTPException(404, "Compra no existe")
    return {
        "id": compra.id, "folio": compra.folio_interno,
        "proveedor_id": compra.proveedor_id,
        "uuid_cfdi": compra.uuid_cfdi,
        "folio_factura_proveedor": compra.folio_factura_proveedor,
        "fecha_recepcion": compra.fecha_recepcion.isoformat(),
        "subtotal": float(compra.subtotal), "iva": float(compra.iva), "total": float(compra.total),
        "estatus": compra.estatus,
        "notas": compra.notas,
        "conceptos": [
            {
                "id": c.id, "variante_id": c.variante_id,
                "descripcion": c.descripcion,
                "cantidad": float(c.cantidad),
                "costo_unitario": float(c.costo_unitario),
                "importe": float(c.importe),
            }
            for c in compra.conceptos
        ],
    }


@router.post("/compras")
def crear_compra(
    payload: CompraIn,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    try:
        compra = compra_service.crear_compra(db, payload, empresa_id)
        return {"id": compra.id, "folio": compra.folio_interno, "total": float(compra.total)}
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/abono")
def registrar_abono(
    payload: AbonoCxPIn,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    try:
        a = compra_service.aplicar_abono_cxp(db, payload, empresa_id)
        return {"id": a.id, "monto": float(a.monto)}
    except ValueError as e:
        raise HTTPException(400, str(e))


# ===== Export XLSX del tablero CxP =====

@router.get("/cartera-xlsx")
def export_cartera_proveedores_xlsx(
    anio: int | None = None,
    mes: int | None = None,
    incluir_pagadas: bool = False,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """Export imprimible para entregar a familia/equipo: a quien debemos, cuanto, cuando vence."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import Response

    q = (
        db.query(CuentaPorPagar, Proveedor)
        .join(Proveedor, Proveedor.id == CuentaPorPagar.proveedor_id)
        .outerjoin(Compra, Compra.id == CuentaPorPagar.compra_id)
        .filter(or_(
            CuentaPorPagar.empresa_id == empresa_id,
            Compra.empresa_id == empresa_id,
        ))
    )
    if not incluir_pagadas:
        q = q.filter(CuentaPorPagar.pagado == False)
    if anio and mes:
        ini = datetime(anio, mes, 1)
        fin = datetime(anio + 1, 1, 1) if mes == 12 else datetime(anio, mes + 1, 1)
        q = q.filter(or_(
            CuentaPorPagar.fecha_vencimiento.between(ini, fin),
            CuentaPorPagar.fecha_recepcion.between(ini, fin),
        ))

    filas = q.all()
    # Ordenar por fecha vencimiento
    filas.sort(key=lambda r: (
        r[0].fecha_vencimiento or r[0].fecha_recepcion or datetime.max
    ))

    wb = Workbook()
    ws = wb.active
    ws.title = "CxP"

    titulo = "Cuentas por Pagar"
    if anio and mes:
        meses = ["", "Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
        titulo += f" — {meses[mes]} {anio}"

    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    ws["A2"] = f"Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="6B7280")
    ws.merge_cells("A2:H2")

    headers = ["Folio", "Fecha llegada", "Fecha vence", "Empresa",
               "Obs", "Monto", "Saldado", "Saldo"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.alignment = Alignment(horizontal="center")

    row = 5
    total_monto = 0.0
    total_saldado = 0.0
    total_saldo = 0.0
    for cxp, prov in filas:
        compra = db.get(Compra, cxp.compra_id) if cxp.compra_id else None
        folio = cxp.folio_factura or (compra.folio_interno if compra else "")
        ws.cell(row=row, column=1, value=folio)
        ws.cell(row=row, column=2, value=cxp.fecha_recepcion.strftime("%d/%m/%Y") if cxp.fecha_recepcion else "")
        ws.cell(row=row, column=3, value=cxp.fecha_vencimiento.strftime("%d/%m/%Y") if cxp.fecha_vencimiento else "")
        ws.cell(row=row, column=4, value=prov.nombre)
        ws.cell(row=row, column=5, value=cxp.observaciones or "")
        m = float(cxp.monto_original)
        s = float(cxp.saldo)
        ws.cell(row=row, column=6, value=m).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=7, value=m - s).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=8, value=s).number_format = '"$"#,##0.00'
        total_monto += m
        total_saldado += (m - s)
        total_saldo += s
        # Resaltar si esta vencida
        if cxp.fecha_vencimiento and cxp.fecha_vencimiento.date() < datetime.utcnow().date() and not cxp.pagado:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FEE2E2")
        row += 1

    # Totales
    row += 1
    ws.cell(row=row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_monto).number_format = '"$"#,##0.00'
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_saldado).number_format = '"$"#,##0.00'
    ws.cell(row=row, column=7).font = Font(bold=True)
    ws.cell(row=row, column=8, value=total_saldo).number_format = '"$"#,##0.00'
    ws.cell(row=row, column=8).font = Font(bold=True)
    ws.cell(row=row, column=8).fill = PatternFill("solid", fgColor="DBEAFE")

    for col, w in zip("ABCDEFGH", [14, 12, 12, 26, 30, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    suffix = f"{anio}-{mes:02d}" if anio and mes else "todas"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="cxp_{suffix}.xlsx"'},
    )


@router.get("/movimientos-xlsx")
def export_movimientos_periodo_xlsx(
    desde: str = Query(..., description="Fecha inicio YYYY-MM-DD"),
    hasta: str = Query(..., description="Fecha fin YYYY-MM-DD (inclusiva)"),
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    """Export XLSX con todos los movimientos CxP en un rango de fechas.
    3 hojas: CxP creadas, Abonos pagados, Resumen por proveedor."""
    from io import BytesIO
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import Response

    try:
        ini = _dt.fromisoformat(desde)
        fin = _dt.fromisoformat(hasta).replace(hour=23, minute=59, second=59)
    except (TypeError, ValueError):
        raise HTTPException(400, "Formato de fechas debe ser YYYY-MM-DD")

    # Hoja 1: CxP creadas en el periodo (fecha_recepcion en rango)
    cxps_creadas = (
        db.query(CuentaPorPagar, Proveedor)
        .join(Proveedor, Proveedor.id == CuentaPorPagar.proveedor_id)
        .outerjoin(Compra, Compra.id == CuentaPorPagar.compra_id)
        .filter(or_(
            CuentaPorPagar.empresa_id == empresa_id,
            Compra.empresa_id == empresa_id,
        ))
        .filter(CuentaPorPagar.fecha_recepcion.between(ini, fin))
        .order_by(CuentaPorPagar.fecha_recepcion)
        .all()
    )

    # Hoja 2: Abonos hechos en el periodo
    abonos = (
        db.query(AbonoCxP, CuentaPorPagar, Proveedor)
        .join(CuentaPorPagar, CuentaPorPagar.id == AbonoCxP.cxp_id)
        .join(Proveedor, Proveedor.id == CuentaPorPagar.proveedor_id)
        .outerjoin(Compra, Compra.id == CuentaPorPagar.compra_id)
        .filter(or_(
            CuentaPorPagar.empresa_id == empresa_id,
            Compra.empresa_id == empresa_id,
        ))
        .filter(AbonoCxP.fecha.between(ini, fin))
        .order_by(AbonoCxP.fecha)
        .all()
    )

    wb = Workbook()

    # === HOJA 1: CxP creadas ===
    ws1 = wb.active
    ws1.title = "CxP del periodo"
    titulo = f"CxP recibidas — {ini.date()} a {fin.date()}"
    ws1["A1"] = titulo
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:H1")
    ws1["A2"] = f"Generado: {_dt.utcnow().strftime('%d/%m/%Y %H:%M')}"
    ws1["A2"].font = Font(italic=True, size=10, color="6B7280")
    ws1.merge_cells("A2:H2")

    headers1 = ["Fecha recepción", "Fecha vence", "Proveedor", "Folio",
                "Obs", "Monto original", "Saldado", "Saldo actual"]
    for col, h in enumerate(headers1, start=1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.alignment = Alignment(horizontal="center")

    row = 5
    total_monto = total_saldado = total_saldo = 0.0
    for cxp, prov in cxps_creadas:
        compra = db.get(Compra, cxp.compra_id) if cxp.compra_id else None
        folio = cxp.folio_factura or (compra.folio_interno if compra else "")
        m = float(cxp.monto_original or 0)
        s = float(cxp.saldo or 0)
        ws1.cell(row=row, column=1, value=cxp.fecha_recepcion.strftime("%d/%m/%Y") if cxp.fecha_recepcion else "")
        ws1.cell(row=row, column=2, value=cxp.fecha_vencimiento.strftime("%d/%m/%Y") if cxp.fecha_vencimiento else "")
        ws1.cell(row=row, column=3, value=prov.nombre)
        ws1.cell(row=row, column=4, value=folio)
        ws1.cell(row=row, column=5, value=cxp.observaciones or "")
        ws1.cell(row=row, column=6, value=m).number_format = '"$"#,##0.00'
        ws1.cell(row=row, column=7, value=m - s).number_format = '"$"#,##0.00'
        ws1.cell(row=row, column=8, value=s).number_format = '"$"#,##0.00'
        total_monto += m
        total_saldado += (m - s)
        total_saldo += s
        row += 1
    row += 1
    ws1.cell(row=row, column=5, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=row, column=6, value=total_monto).number_format = '"$"#,##0.00'
    ws1.cell(row=row, column=6).font = Font(bold=True)
    ws1.cell(row=row, column=7, value=total_saldado).number_format = '"$"#,##0.00'
    ws1.cell(row=row, column=7).font = Font(bold=True)
    ws1.cell(row=row, column=8, value=total_saldo).number_format = '"$"#,##0.00'
    ws1.cell(row=row, column=8).font = Font(bold=True)
    for col, w in zip("ABCDEFGH", [14, 12, 28, 16, 30, 14, 14, 14]):
        ws1.column_dimensions[col].width = w

    # === HOJA 2: Abonos del periodo ===
    ws2 = wb.create_sheet("Abonos del periodo")
    ws2["A1"] = f"Abonos pagados — {ini.date()} a {fin.date()}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:G1")

    headers2 = ["Fecha pago", "Proveedor", "Folio CxP", "Forma pago",
                "Referencia", "Monto pagado", "Notas"]
    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="065F46")
        c.alignment = Alignment(horizontal="center")

    row = 4
    total_abonos = 0.0
    for ab, cxp, prov in abonos:
        compra = db.get(Compra, cxp.compra_id) if cxp.compra_id else None
        folio = cxp.folio_factura or (compra.folio_interno if compra else "")
        monto_ab = float(ab.monto or 0)
        ws2.cell(row=row, column=1, value=ab.fecha.strftime("%d/%m/%Y") if ab.fecha else "")
        ws2.cell(row=row, column=2, value=prov.nombre)
        ws2.cell(row=row, column=3, value=folio)
        ws2.cell(row=row, column=4, value=ab.forma_pago or "")
        ws2.cell(row=row, column=5, value=ab.referencia or "")
        ws2.cell(row=row, column=6, value=monto_ab).number_format = '"$"#,##0.00'
        ws2.cell(row=row, column=7, value=ab.notas or "")
        total_abonos += monto_ab
        row += 1
    row += 1
    ws2.cell(row=row, column=5, value="TOTAL ABONOS").font = Font(bold=True)
    ws2.cell(row=row, column=6, value=total_abonos).number_format = '"$"#,##0.00'
    ws2.cell(row=row, column=6).font = Font(bold=True)
    ws2.cell(row=row, column=6).fill = PatternFill("solid", fgColor="DCFCE7")
    for col, w in zip("ABCDEFG", [14, 28, 16, 14, 18, 14, 30]):
        ws2.column_dimensions[col].width = w

    # === HOJA 3: Resumen por proveedor ===
    ws3 = wb.create_sheet("Resumen proveedor")
    ws3["A1"] = f"Resumen por proveedor — {ini.date()} a {fin.date()}"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.merge_cells("A1:E1")

    # Agregar: por proveedor, total CxP recibido en periodo + total abonado en periodo
    resumen: dict[int, dict] = {}
    for cxp, prov in cxps_creadas:
        r = resumen.setdefault(prov.id, {"nombre": prov.nombre, "recibido": 0.0, "pagado": 0.0, "facturas": 0, "abonos": 0})
        r["recibido"] += float(cxp.monto_original or 0)
        r["facturas"] += 1
    for ab, _cxp, prov in abonos:
        r = resumen.setdefault(prov.id, {"nombre": prov.nombre, "recibido": 0.0, "pagado": 0.0, "facturas": 0, "abonos": 0})
        r["pagado"] += float(ab.monto or 0)
        r["abonos"] += 1

    headers3 = ["Proveedor", "Facturas recibidas", "Monto recibido",
                "Abonos hechos", "Monto pagado"]
    for col, h in enumerate(headers3, start=1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E40AF")
        c.alignment = Alignment(horizontal="center")

    row = 4
    tot_rec = tot_pag = 0.0
    for _pid, r in sorted(resumen.items(), key=lambda x: x[1]["recibido"], reverse=True):
        ws3.cell(row=row, column=1, value=r["nombre"])
        ws3.cell(row=row, column=2, value=r["facturas"])
        ws3.cell(row=row, column=3, value=r["recibido"]).number_format = '"$"#,##0.00'
        ws3.cell(row=row, column=4, value=r["abonos"])
        ws3.cell(row=row, column=5, value=r["pagado"]).number_format = '"$"#,##0.00'
        tot_rec += r["recibido"]
        tot_pag += r["pagado"]
        row += 1
    row += 1
    ws3.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws3.cell(row=row, column=3, value=tot_rec).number_format = '"$"#,##0.00'
    ws3.cell(row=row, column=3).font = Font(bold=True)
    ws3.cell(row=row, column=5, value=tot_pag).number_format = '"$"#,##0.00'
    ws3.cell(row=row, column=5).font = Font(bold=True)
    for col, w in zip("ABCDE", [32, 16, 16, 16, 16]):
        ws3.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f"cxp_movimientos_{ini.date()}_a_{fin.date()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ===== Deuda bancaria con conceptos editables =====

@router.get("/deuda-bancaria")
def listar_deuda_bancaria(
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from app.models import DeudaBancaria
    deudas = (
        db.query(DeudaBancaria)
        .filter(DeudaBancaria.empresa_id == empresa_id)
        .filter(DeudaBancaria.activa == True)
        .order_by(DeudaBancaria.id)
        .all()
    )
    out = []
    for d in deudas:
        total = sum(float(c.monto) for c in d.conceptos)
        out.append({
            "id": d.id, "nombre": d.nombre, "referencia": d.referencia,
            "notas": d.notas, "total": round(total, 2),
            "conceptos": [
                {
                    "id": c.id, "concepto": c.concepto,
                    "monto": float(c.monto), "orden": c.orden,
                    "pct": (float(c.monto) / total * 100) if total > 0 else 0,
                }
                for c in d.conceptos
            ],
        })
    return out


class DeudaBancariaIn(BaseModel):
    nombre: str
    referencia: str | None = None
    notas: str | None = None


@router.post("/deuda-bancaria")
def crear_deuda_bancaria(
    payload: DeudaBancariaIn,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from app.models import DeudaBancaria
    d = DeudaBancaria(
        empresa_id=empresa_id,
        nombre=payload.nombre.strip(),
        referencia=(payload.referencia or "").strip() or None,
        notas=payload.notas,
    )
    db.add(d)
    db.commit()
    db.refresh(d)
    return {"id": d.id}


@router.patch("/deuda-bancaria/{deuda_id}")
def actualizar_deuda_bancaria(
    deuda_id: int,
    payload: dict,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from app.models import DeudaBancaria
    d = db.get(DeudaBancaria, deuda_id)
    if not d or d.empresa_id != empresa_id:
        raise HTTPException(404, "Deuda no existe")
    for k in ("nombre", "referencia", "notas"):
        if k in payload:
            setattr(d, k, payload[k] or None if k != "nombre" else payload[k])
    db.commit()
    return {"ok": True}


@router.delete("/deuda-bancaria/{deuda_id}")
def borrar_deuda_bancaria(
    deuda_id: int,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from app.models import DeudaBancaria
    d = db.get(DeudaBancaria, deuda_id)
    if not d or d.empresa_id != empresa_id:
        raise HTTPException(404, "Deuda no existe")
    d.activa = False
    db.commit()
    return {"ok": True}


class ConceptoDeudaIn(BaseModel):
    concepto: str
    monto: float = 0


@router.post("/deuda-bancaria/{deuda_id}/conceptos")
def crear_concepto_deuda(
    deuda_id: int,
    payload: ConceptoDeudaIn,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from app.models import DeudaBancaria, ConceptoDeudaBancaria
    d = db.get(DeudaBancaria, deuda_id)
    if not d or d.empresa_id != empresa_id:
        raise HTTPException(404, "Deuda no existe")
    orden = (db.query(func.coalesce(func.max(ConceptoDeudaBancaria.orden), 0))
             .filter(ConceptoDeudaBancaria.deuda_id == deuda_id).scalar() or 0) + 1
    c = ConceptoDeudaBancaria(
        deuda_id=deuda_id, concepto=payload.concepto.strip(),
        monto=payload.monto, orden=orden,
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return {"id": c.id}


@router.patch("/deuda-bancaria/conceptos/{cid}")
def actualizar_concepto_deuda(
    cid: int,
    payload: dict,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from app.models import ConceptoDeudaBancaria, DeudaBancaria
    c = db.get(ConceptoDeudaBancaria, cid)
    if not c:
        raise HTTPException(404, "Concepto no existe")
    d = db.get(DeudaBancaria, c.deuda_id)
    if d.empresa_id != empresa_id:
        raise HTTPException(403, "Deuda de otra empresa")
    if "concepto" in payload:
        c.concepto = payload["concepto"]
    if "monto" in payload:
        c.monto = float(payload["monto"])
    db.commit()
    return {"ok": True}


@router.delete("/deuda-bancaria/conceptos/{cid}")
def borrar_concepto_deuda(
    cid: int,
    empresa_id: int = Depends(get_active_empresa_id),
    db: Session = Depends(get_db),
):
    from app.models import ConceptoDeudaBancaria, DeudaBancaria
    c = db.get(ConceptoDeudaBancaria, cid)
    if not c:
        raise HTTPException(404, "Concepto no existe")
    d = db.get(DeudaBancaria, c.deuda_id)
    if d.empresa_id != empresa_id:
        raise HTTPException(403, "Deuda de otra empresa")
    db.delete(c)
    db.commit()
    return {"ok": True}
